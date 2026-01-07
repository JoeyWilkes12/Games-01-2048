import openpyxl
import json
import os
from datetime import datetime

input_file = '2048 Robots.xlsx'
output_file = 'dashboard_data.json'

def extract_data():
    try:
        wb = openpyxl.load_workbook(input_file, data_only=True)
        
        # 1. Process 'Record' sheet
        record_sheet = wb['Record']
        records = []
        
        # Get headers from first row
        headers = [cell.value for cell in record_sheet[1]]
        
        # Iterate rows up to 149 (1-based index in openpyxl, user said data valid to row 149)
        # Note: sheet rows are 1-based. user said "row 149", likely including header.
        # Let's iterate from 2 to 149.
        for row in record_sheet.iter_rows(min_row=2, max_row=149, values_only=True):
            record = {}
            has_data = False
            for i, value in enumerate(row):
                if i < len(headers):
                    header = headers[i]
                    # Handle potential None headers
                    if header:
                        record[str(header)] = value
                        if value is not None:
                            has_data = True
            
            if has_data:
                records.append(record)
                
        # 2. Process 'Legend' sheet
        legend_sheet = wb['Legend']
        legends = []
        legend_headers = [cell.value for cell in legend_sheet[1]]
        
        for row in legend_sheet.iter_rows(min_row=2, values_only=True):
            legend = {}
            has_data = False
            for i, value in enumerate(row):
                if i < len(legend_headers):
                    header = legend_headers[i]
                    if header:
                        legend[str(header)] = value
                        if value is not None:
                            has_data = True
            if has_data:
                legends.append(legend)

        data = {
            'records': records,
            'legends': legends,
            'generatedAt': datetime.now().isoformat()
        }
        
        output_js_file = 'dashboard_data.js'
        with open(output_js_file, 'w') as f:
            json_str = json.dumps(data, indent=2, default=str)
            f.write(f"const dashboardData = {json_str};")
            
        print(f"Successfully extracted data to {output_js_file}")
        print(f"Extracted {len(records)} records and {len(legends)} legends.")
        
    except Exception as e:
        print(f"Error extracting data: {e}")
        exit(1)

if __name__ == "__main__":
    extract_data()
